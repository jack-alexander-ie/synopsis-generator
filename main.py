#!/usr/bin/env python3
# -*- coding: latin-1 -*-

import os
import sys
import textwrap

# viewing tidied array
from pprint import pprint

try:
    from PIL import Image, ImageFont, ImageDraw, ImageOps
    from openpyxl import load_workbook
    import numpy as np
    from googletrans import Translator

except ImportError:
    print("Some packages not installed.")
    print("Script uses Python 3.x, please install using: pip3 install [--user] [package name]")
    sys.exit(1)

# Global Folder Paths
FOLDER_PATH = os.getcwd()
BANNER_PATH = "source/banners/%s.png"
COVERS_PATH = "source/covers/%s.png"
SAVED_PATH = "images/%s/%s Synopsis %s.png"

# FONT SETTINGS #

font_file_reg = 'source/fonts/texgyreheros-regular.otf'
font_file_bold = 'source/fonts/texgyreheros-bold.otf'

title_font_size, synopsis_font_size, text_font_size, sub_title_font_size, movie_title_font_size = 30, 14, 12, 20, 13

artist_font_size, album_font_size = 28, 22

font_title_reg = ImageFont.truetype(font_file_reg, title_font_size, encoding="unic")
font_title_text = ImageFont.truetype(font_file_reg, text_font_size, encoding="unic")
font_title_syn = ImageFont.truetype(font_file_reg, synopsis_font_size, encoding="unic")
font_sub_title = ImageFont.truetype(font_file_reg, sub_title_font_size, encoding="unic")

font_album_bold = ImageFont.truetype(font_file_reg, artist_font_size, encoding="unic")
font_album_reg = ImageFont.truetype(font_file_reg, album_font_size, encoding="unic")

font_movie_cover = ImageFont.truetype(font_file_bold, movie_title_font_size, encoding="unic")

white = 255, 255, 255, 255
grey = 160, 160, 160, 255
dark_grey = 32, 32, 32
line_grey = 45, 45, 45

# GLOBAL VARIABLES #

video_media = ['movies', 'tv', 'kids']

kids_synopsis_width = 95


def convert_image_to_png(path):
    """

    Helper function to convert a jpeg to a png.

    :param path: Path to image to be converted
    :return: Saves the converted image to the root folder

    """

    files = []

    for file in os.listdir(path):

        if file.endswith(".jpg") or file.endswith(".jpeg"):
            files.append(file)

    for image in files:

        image_original = Image.open(path + image)
        new_image_name = ''

        if image.endswith(".jpg"):

            new_image_name = image[:-4]

        elif image.endswith(".jpeg"):

            new_image_name = image[:-5]

        image_original.save(path + new_image_name + '.png')


def verify_path(file_path):
    # TODO: Verify file path validity check

    try:
        loaded_image = Image.open(file_path).convert("RGBA")
        return loaded_image
    except LookupError:
        print("File does not exist. Please check the file path")


class LanguageTranslator(object):
    supported_languages = {
        "greek": ["el", "ell"], "esperanto": ["eo", "epo"], "english": ["en", "eng"],
        "afrikaans": ["af", "afr"], "swahili": ["sw", "swa"], "catalan": ["ca", "cat"],
        "italian": ["it", "ita"], "hebrew": ["iw", "heb"], "swedish": ["sv", "swe"],
        "czech": ["cs", "cze"], "welsh": ["cy", "cym"], "arabic": ["ar", "ara"],
        "urdu": ["ur", "urd"], "irish": ["ga", "gle"], "basque": ["eu", "eus"],
        "estonian": ["et", "est"], "azerbaijani": ["az", "aze"], "indonesian": ["id", "ind"],
        "spanish": ["es", "spa"], "russian": ["ru", "rus"], "galician": ["gl", "glg"],
        "dutch": ["nl", "nld"], "portuguese": ["pt", "por"], "latin": ["la", "lat"],
        "turkish": ["tr", "tur"], "filipino": ["tl", "fil"], "latvian": ["lv", "lav"],
        "lithuanian": ["lt", "lit"], "thai": ["th", "tha"], "vietnamese": ["vi", "vie"],
        "gujarati": ["gu", "guj"], "romanian": ["ro", "ron"], "icelandic": ["is", "isl"],
        "polish": ["pl", "pol"], "tamil": ["ta", "tam"], "yiddish": ["yi", "yid"],
        "belarusian": ["be", "bel"], "french": ["fr", "fre"], "bulgarian": ["bg", "bul"],
        "ukrainian": ["uk", "ukr"], "croatian": ["hr", "hrv"], "bengali": ["bn", "ben"],
        "slovenian": ["sl", "slv"], "haitian-creole": ["ht", "hat"], "danish": ["da", "dan"],
        "persian": ["fa", "fas"], "hindi": ["hi", "hin"], "finnish": ["fi", "fin"],
        "hungarian": ["hu", "hun"], "japanese": ["ja", "jpn"], "georgian": ["ka", "kat"],
        "telugu": ["te", "tel"], "chinese-traditional": ["zh-TW", "chi"], "albanian": ["sq", "sqi"],
        "norwegian": ["no", "nor"], "korean": ["ko", "kor"], "kannada": ["kn", "kan"],
        "macedonian": ["mk", "mkd"], "chinese-simplified": ["zh-CN", "chi"], "slovak": ["sk", "slk"],
        "maltese": ["mt", "mlt"], "german": ["de", "deu"], "malay": ["ms", "msa"],
        "serbian": ["sr", "srp"]
    }

    custom_languages = {
        "greenlandic": ["kl", "kal"]
    }

    def iso_checker(self, language):
        """
        Returns relevant ISO codes for the selected language

        :param language: language entered by user
        :return: array with 2 letter and 3 letter ISO
        """

        language = language.lower()

        if language in self.supported_languages:

            iso_codes = self.supported_languages[language][0], self.supported_languages[language][1].upper(), True

        else:
            print(language, 'not in supported languages - looking in custom support list')

            if language in self.custom_languages:
                iso_codes = self.custom_languages[language][0], self.custom_languages[language][1].upper(), False

        return iso_codes

    @staticmethod
    def translate_word(word, two_letter_iso):

        print("Translation needed for '" + word + "', calling Google...")

        translator = Translator()

        translated = ''

        try:
            translation = translator.translate(word, dest=two_letter_iso)
            translated = translation.text
        except ConnectionError:
            print('Connection Error, try again.')

        return translated

    @staticmethod
    def custom_translation(word, language):

        custom_translations = {

            "KAL": {
                "Duration": "Sivisussusaa",
                "Director": "Illitsersuisoq",
                "Language": "Oqaatsit",
                "Rating": "Nalilersuineq",
                "Starring": "Peqataasut",
                "minutes": "minutsi"
            }
        }

        translated_word = ""

        for translation in custom_translations[language]:
            if word in translation:
                translated_word = custom_translations[language][word]

        return translated_word


class DataParser(object):

    def __init__(self):

        self.prepared_data = {}

    @staticmethod
    def retrieve_excel_data(workbook_sheet):
        """
        Scrapes all data from sheet in DATA.xlsx and sorts it based on the column title.

        :param workbook_sheet: name of the sheet to scrape
        :return: list of dicts containing all the data
        """

        DATASHEET = FOLDER_PATH + "/source/spreadsheet/DATA.xlsx"

        wb = load_workbook(filename=DATASHEET, data_only=True, read_only=True)
        type(wb)

        sheets = [workbook_sheet, workbook_sheet + "_DIMENSIONS"]

        sheet_data = []

        sheet_identifier = 0

        for sheet in sheets:

            # Get film data
            sheet = wb[sheet]

            headers = []
            data = []

            # Get header values
            for row in sheet.iter_rows(max_row=1):
                for cell in row:
                    if cell.value is not None:
                        headers.append(cell.value)

            # Determine first empty cell
            max_row = 0

            for row in sheet.iter_rows(row_offset=1):
                value = sheet['A' + str(max_row + 1)].value
                if value is not None:
                    max_row += 1

            for row in sheet.iter_rows(row_offset=1, max_row=max_row):

                metadata = {}
                count = 0

                while count < len(headers):
                    metadata[headers[count]] = row[count].value
                    count += 1

                # Remove some unnecessary values
                if sheet_identifier == 0:
                    metadata['TYPE'] = workbook_sheet
                else:
                    metadata.pop('SPACING')

                data.append(metadata)

            sheet_data.append(data)

            sheet_identifier += 1

        return sheet_data

    def prepare_movie_data(self, iso_codes, workbook_sheet):
        """
        Prepares the data so that it can be drawn more efficiently.

        :param workbook_sheet: sheet from which to scrape data
        :param iso_codes: array with relevant iso codes for translations and file naming

        :return: returns an array of dicts containing the prepared data

        """

        film_metadata, raw_coordinates = self.retrieve_excel_data(workbook_sheet)

        # Required for Google Translate
        two_letter_iso, three_letter_iso, supported_flag = iso_codes

        coordinates = {}

        minutes = "minutes"

        if three_letter_iso not in "ENG":

            ''' If the language is not English '''

            translator = LanguageTranslator()

            if supported_flag is True:

                ''' If the language is supported by Google Translate '''

                for coordinate in raw_coordinates:

                    if coordinate['VALUE'] is not None:

                        word = coordinate['VALUE']

                        try:
                            translated_word = translator.translate_word(word, two_letter_iso)
                            coordinate['VALUE'] = translated_word

                        except ValueError:
                            print('Error with value:', three_letter_iso)

                    coordinates[coordinate['KEY']] = coordinate

                    # KEY no longer needed..
                    coordinate.pop('KEY')

                minutes = translator.translate_word(minutes, two_letter_iso)

            elif supported_flag is False:

                ''' If the language is not supported by Google Translate '''

                for coordinate in raw_coordinates:

                    if coordinate['VALUE'] is not None:
                        word = coordinate['VALUE']

                        translated_word = translator.custom_translation(word, three_letter_iso)
                        coordinate['VALUE'] = translated_word

                    coordinates[coordinate['KEY']] = coordinate

                    # KEY no longer needed..
                    coordinate.pop('KEY')

                minutes = translator.custom_translation(minutes, three_letter_iso)

        else:

            ''' If the language is English '''

            for coordinate in raw_coordinates:
                coordinates[coordinate['KEY']] = coordinate

                # KEY no longer needed..
                coordinate.pop('KEY')

        films = []

        # Package film data
        for film in film_metadata:

            cover_path = COVERS_PATH % (film['TITLE'],)
            banner_path = BANNER_PATH % (film['TITLE'],)
            save_path = SAVED_PATH % (film['TYPE'].lower(), film['TITLE'], iso_codes[1])

            media_info = {
                "Type": film['TYPE'],
                "Name": film['TITLE'],
                "Paths": {"Cover": cover_path, "Banner": banner_path, "Save": save_path},

                "Duration": {"Header": coordinates['Duration']['VALUE'],
                             "Header_X": coordinates['Duration']['HEADER_X_POS'],
                             "Header_Y": coordinates['Duration']['HEADER_Y_POS'],
                             "Info": str(film['DURATION']) + " " + minutes,
                             "Info_X": coordinates['Duration']['INFO_X_POS'],
                             "Info_Y": coordinates['Duration']['INFO_Y_POS']},

                "Language": {"Header": coordinates['Language']['VALUE'],
                             "Header_X": coordinates['Language']['HEADER_X_POS'],
                             "Header_Y": coordinates['Language']['HEADER_Y_POS'],
                             "Info": film['LANGUAGE'],
                             "Info_X": coordinates['Language']['INFO_X_POS'],
                             "Info_Y": coordinates['Language']['INFO_Y_POS']},

                "Rating": {"Header": coordinates['Rating']['VALUE'],
                           "Header_X": coordinates['Rating']['HEADER_X_POS'],
                           "Header_Y": coordinates['Rating']['HEADER_Y_POS'],
                           "Info": film['RATING'],
                           "Info_X": coordinates['Rating']['INFO_X_POS'],
                           "Info_Y": coordinates['Rating']['INFO_Y_POS']},

                "Starring": {"Header": coordinates['Starring']['VALUE'],
                             "Header_X": coordinates['Starring']['HEADER_X_POS'],
                             "Header_Y": coordinates['Starring']['HEADER_Y_POS'],
                             "Info": textwrap.wrap(film['STARRING'], width=coordinates['Starring']['WIDTH']),
                             "Info_X": coordinates['Starring']['INFO_X_POS'],
                             "Info_Y": coordinates['Starring']['INFO_Y_POS']},

                "Synopsis": {
                    "Info": textwrap.wrap(film["SYNOPSIS_" + three_letter_iso], width=coordinates['Synopsis']['WIDTH']),
                    "Info_X": coordinates['Synopsis']['INFO_X_POS'],
                    "Info_Y": coordinates['Synopsis']['INFO_Y_POS']},

                "Title": {"Info": film["TITLE_" + three_letter_iso],
                          "Info_X": coordinates['Title']['INFO_X_POS'],
                          "Info_Y": coordinates['Title']['INFO_Y_POS']}
            }

            # TODO: Add extra information - episode, season etc

            # Remove data not needed for other media types
            if media_info['Type'] in 'MOVIES':

                media_info["Director"] = {"Header": coordinates['Director']['VALUE'],
                                          "Header_X": coordinates['Director']['HEADER_X_POS'],
                                          "Header_Y": coordinates['Director']['HEADER_Y_POS'],
                                          "Info": film['DIRECTOR'],
                                          "Info_X": coordinates['Director']['INFO_X_POS'],
                                          "Info_Y": coordinates['Director']['INFO_Y_POS']}

            elif media_info['Type'] in 'TV':

                print()

            elif media_info['Type'] in 'KIDS':

                removables = ['Director', 'Language', 'Rating', 'Starring']

                for removable in removables:
                    media_info.pop(removable)

            films.append(media_info)

        # Save memory
        del film_metadata, raw_coordinates

        return films


class GenerateBlankTemplate(object):
    """

    Generates a blank synopsis panel.

    """

    def __init__(self):
        # TODO: Add blank panel dimensions to the spreadsheet - no goose hunts
        # Generate blank panel on which to draw
        panel_width, panel_height = 698, 565

        # Create blank panel
        blank_panel = Image.new('RGBA', (panel_width, panel_height), dark_grey)

        # Grey line at the bottom of the screen
        self.draw_grey_line(blank_panel)

    @staticmethod
    def draw_grey_line(merged_image):
        # Grey line at the bottom of the panel
        draw = ImageDraw.Draw(merged_image)

        grey_line_y_pos = 466
        grey_line_x_pos = 27
        grey_line_length = 408 + grey_line_x_pos

        draw.line((grey_line_x_pos, grey_line_y_pos) + (grey_line_length, grey_line_y_pos), fill=line_grey)

        # Save to disk
        merged_image.save('source/blank_panel.png')


class MetadataImages(object):

    def __init__(self, mt, iso_codes):

        parsed_data = DataParser()

        films = parsed_data.prepare_movie_data(iso_codes, mt)

        # Checks if blank panel ready to go - if not, it creates one
        self.blank_template = self.blank_panel_check()

        # Create images for each film in array
        for film in films:
           self.create_synopsis_images(film)
           create_cover_image = Covers(film)

    def create_synopsis_images(self, film):

        # Create banner image
        banner_panel = self.create_banner_template(film['Paths']['Banner'])

        # Merge blank and banner panels
        template_image = self.merge_images(self.blank_template, banner_panel)

        # Prepare it to be drawn on
        draw = ImageDraw.Draw(template_image)

        # Draw info to the image
        self.draw_info(draw, film)

        if film['Type'] is 'KIDS':
            kids_icon = Image.open('source/icons/time.png').convert('RGBA')

            # TODO: Add paste dimensions to the spreadsheet - no goose hunts
            # Paste on Kids clock icon
            template_image.paste(kids_icon, (27, 400), kids_icon)

        # Save to disk
        template_image.save(film['Paths']['Save'])

    def blank_panel_check(self):

        exists = os.path.isfile('source/blank_panel.png')

        if not exists:
            # Make a blank template
            GenerateBlankTemplate()

        # Load the blank template
        blank_panel = self.load_template()

        return blank_panel

    @staticmethod
    def load_template():

        # Load a blank template to draw on
        blank_template = Image.open('source/blank_panel.png').convert('RGBA')

        return blank_template

    def create_banner_template(self, banner_path):

        blended_banner = self.apply_blend_to_banner(banner_path)

        empty_png = Image.new('RGBA', (698, 565), (255, 255, 255, 0))
        empty_png.paste(blended_banner, (0, 0))

        return empty_png

    @staticmethod
    def apply_blend_to_banner(file_path):

        """

        Applies a blend to the bottom of the banner image.

        :param file_path:
        :return:
        """

        image = verify_path(file_path)

        width, height = image.size

        pixels = image.load()

        for y in range(int(height * .80), int(height * 1)):

            alpha = 255 - int((y - height * .80) / height / .20 * 255)

            for x in range(width):
                pixels[x, y] = pixels[x, y][:3] + (alpha,)

        for y in range(y, height):

            for x in range(width):
                pixels[x, y] = pixels[x, y][:3] + (0,)

        return image

    @staticmethod
    def merge_images(blank_panel, banner_panel):

        merged_image = Image.new('RGBA', blank_panel.size)
        merged_image = Image.alpha_composite(merged_image, blank_panel)
        merged_image = Image.alpha_composite(merged_image, banner_panel)

        return merged_image

    @staticmethod
    def draw_info(draw, media_data):

        keys_to_ignore = ['Paths', 'Synopsis', 'Title', 'Type', 'Name']
        standard_draw = ['Duration', 'Rating', 'Language', 'Director']

        for key in media_data:

            # Draw header
            if key not in keys_to_ignore:
                draw.text((media_data[key]['Header_X'], media_data[key]['Header_Y']), media_data[key]['Header'],
                          font=font_title_text, fill=white)

            # Draw info based on key
            if key in standard_draw:

                draw.text((media_data[key]['Info_X'], media_data[key]['Info_Y']), media_data[key]['Info'],
                          font=font_title_text, fill=grey)

            elif key is 'Starring':

                text_offset = 490

                for line in media_data[key]['Info']:
                    x_pos = media_data[key]['Info_X']
                    y_pos = text_offset + (media_data[key]['Info_Y'] - x_pos)

                    draw.text((x_pos, y_pos), line, font=font_title_text, fill=grey)

                    text_offset += font_title_text.getsize(line)[1]

            elif key is 'Synopsis':

                offset = 27

                for line in media_data[key]['Info']:
                    draw.text(
                        (media_data[key]['Info_X'], offset + (media_data[key]['Info_Y'] - media_data[key]['Info_X'])),
                        line,
                        font=font_title_syn, fill=grey)

                    offset += font_title_syn.getsize(line)[1]

            elif key is 'Title':

                draw.text((media_data[key]['Info_X'], media_data[key]['Info_Y']), media_data[key]['Info'],
                          font=font_title_reg, fill=white)

            elif key is 'Episode':

                draw.text((media_data[key]['Info_X'], media_data[key]['Info_y']), media_data[key]['Info'],
                          font=font_sub_title, fill=white)


class Covers(object):

    def __init__(self, film):

        # TODO: Add cover dimensions to the spreadsheet - no goose hunts
        self.cover_width = 130
        self.cover_height = 187

        self.file_path = film['Paths']['Cover']
        self.cover_title = film['Title']['Info']
        self.output_name = film['Title']['Info'] + '.png'
        self.covers_file_path = 'images/covers/'

        # Create cover image for the film
        self.create_cover_image()

    def create_cover_image(self):

        cover_art = verify_path(self.file_path)

        # Create blank panel and draw title on it
        panel = self.blank_panel_with_title(self.cover_title)

        # Sort the images so they can be combined
        arranged_image = self.arrange_images(cover_art, panel)

        # Crop the image dow to the right size
        cover_image = self.crop_image(arranged_image)

        # Make a thumbnail of the image if it is album artwork
        # cover_image.thumbnail((self.cover_width, self.cover_height), Image.ANTIALIAS)

        cover_image.save(self.covers_file_path + self.output_name)

    def blank_panel_with_title(self, film_title):

        blank_panel = Image.new('RGBA', (self.cover_width, self.cover_height))

        draw = ImageDraw.Draw(blank_panel)

        film_title = self.length_check(film_title)

        draw.text((0, 5), film_title, font=font_title_text, fill=white)

        return blank_panel

    @staticmethod
    def arrange_images(cover_art, panel):

        list_images = [cover_art, panel]

        min_shape = sorted([(np.sum(i.size), i.size) for i in list_images])[0][1]

        images_comb = np.vstack((np.asarray(i.resize(min_shape)) for i in list_images))

        images_comb = Image.fromarray(images_comb)

        return images_comb

    def crop_image(self, arranged_image):

        # Crop new large image down to size
        left, top, right, bottom = 0, 0, self.cover_width, 210

        cropped_image = arranged_image.crop((left, top, right, bottom))

        return cropped_image

    @staticmethod
    def length_check(title):
        """
        Checks if a movie title is too long.

        Returns a shortened version if it is.

        :param title: Film title as input
        :return: Shortened film title as output
        """

        if len(title) > 18:
            title = title[:17] + "..."
            return title
        else:
            return title


def main(arguments):
    """

    Main function.

    Takes CLI arguments as input and generates images accordingly.

    :param arguments: CLI arguments - media type and language if required
    :return: Generated images with associated metadata
    """

    media_type = arguments[1].lower()
    language = arguments[2]

    translator = LanguageTranslator()

    iso_codes = translator.iso_checker(language)

    print("Creating " + media_type + " images...")

    if media_type in video_media:

        MetadataImages(media_type.upper(), iso_codes)

    elif media_type == 'Music':

        # TODO: Make dynamic to include album artwork
        Covers()

    elif media_type == 'Convert':

        convert_image_to_png('')

    print("Complete. Images in the folder 'images/" + media_type + "/'")


if __name__ == "__main__":
    main(sys.argv)
